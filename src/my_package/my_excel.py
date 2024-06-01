import openpyxl


# 读取excel中的数据，输出类型是datas
def read_to_excel(filename, sheet_index=0):
    ars = []
    # 加载 Excel 工作簿
    workbook = openpyxl.load_workbook(filename)
    # 获取指定索引的工作表
    sheet = workbook.worksheets[sheet_index]
    # 读取工作表的所有行，并将每行转换为列表后添加到 ars
    for s in sheet.values:
        ar = list(s)
        ars.append(ar)
    return ars


def write_to_excel_row(data, filename='output.xlsx', sheet_name='Sheet1'):
    try:
        # 尝试加载已存在的工作簿
        workbook = openpyxl.load_workbook(filename)
        # 尝试获取已存在的工作表
        sheet = workbook[sheet_name]
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的工作簿和工作表
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    # 将数据追加到工作表
    sheet.append(data)

    # 保存工作簿到文件
    workbook.save(filename)


def write_to_excel_all(datas, filename='output.xlsx', sheet_name='Sheet1'):
    try:
        # 尝试加载已存在的工作簿
        workbook = openpyxl.load_workbook(filename)
        # 尝试获取已存在的工作表
        sheet = workbook[sheet_name]
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的工作簿和工作表
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    for data in datas:
        # 将数据追加到工作表
        sheet.append(data)

    # 保存工作簿到文件
    workbook.save(filename)


if __name__ == '__main__':
    datas = [
        [1, 2, 3],
        [4, 5, 6]
    ]
    write_to_excel_all(datas)
